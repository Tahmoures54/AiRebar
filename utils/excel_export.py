# utils/excel_export.py
"""
Excel Export – AI Rebar
Exports full rebar schedule, project info, summary, and shape catalog
to a multi‑sheet Excel workbook with auto‑formatting and company branding.
"""

import pandas as pd
import json
import datetime
import logging
import os
from typing import List, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

from config import WEIGHT_COEFFICIENT, DEFAULT_REBAR_GRADE
from shapes.definitions import default_shape_registry
from logic.calculator import calculate_weight
from db.database import DatabaseManager

logger = logging.getLogger(__name__)
db = DatabaseManager()

# ---------- Branding constants ----------
COMPANY_NAME = "AI Rebar"
COMPANY_TAGLINE = "Intelligent Bar Bending Schedule & Cutting Optimization"
COMPANY_WEBSITE = "https://airebar.io"
COMPANY_PHONE = "+98 916 068 4552"
LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")  # relative to utils/

# ---------- Styles ----------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=12, color="333333")
NORMAL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


def export_excel(
    filepath: str,
    project_id: int,
    project_name: str,
    client_name: str,
    summary_data: Optional[List[List]] = None
) -> None:
    """
    Save a professionally formatted Excel workbook with sheets:
        - Bar Bending Schedule (full rebar list, auto‑formatted)
        - Summary (project info + weight summary)
        - Shape Details (catalog of all defined shapes)
        - About (company branding & logo)

    Parameters
    ----------
    filepath : str
        Output .xlsx path.
    project_id : int
    project_name, client_name : str
    summary_data : list of list, optional
        Rows matching [Diameter, Grade, Total Wt (kg), Total Len (m),
        # Stock Bars, Waste (m), Waste (%)].
    """
    # ===================== 1. Bar Bending Schedule =====================
    columns = (
        "ID", "Listofer No.", "Listofer Desc.", "Pos.", "Dia",
        "Grade", "Shape Code", "Shape", "Dimensions (mm)", "Cut Len (mm)",
        "Qty", "Unit Wt (kg/m)", "Total Wt (kg)",
        "Location", "Element", "Added By", "Date Added"
    )

    query = """
        SELECT r.id, l.number, l.description, r.pos, r.diameter,
               r.shape_name, r.dimensions, r.quantity,
               r.location, r.element_type, r.added_by, r.date_added,
               r.grade
        FROM rebars r
        JOIN listofers l ON r.listofer_id = l.id
        WHERE l.project_id = ?
        ORDER BY l.number, r.pos
    """
    rows = db.fetchall(query, (project_id,))

    data = []
    for row in rows:
        (rid, lnum, ldesc, pos, dia, sname, dims_str,
         qty, loc, etype, added_by, date_added, grade) = row

        if grade is None:
            grade = DEFAULT_REBAR_GRADE

        shape_def = default_shape_registry.get_shape_def(sname)
        if shape_def is None:
            logger.warning("Unknown shape '%s' for rebar %d, skipping.", sname, rid)
            continue

        try:
            if dims_str and dims_str.startswith('{'):
                dims = json.loads(dims_str)
            else:
                dims = {}
                if dims_str:
                    for part in dims_str.split(','):
                        if '=' in part:
                            k, v = part.split('=')
                            dims[k.strip()] = float(v.strip())
        except Exception:
            dims = {}

        try:
            cut_len = default_shape_registry.calc_shape_length(sname, dims, dia)
        except Exception:
            logger.warning("Could not compute length for rebar %d, shape '%s'", rid, sname)
            cut_len = 0.0

        _, unit_wt = calculate_weight(dia, cut_len)
        total_wt = unit_wt * qty

        data.append((
            rid, lnum, ldesc, pos, dia,
            grade,
            shape_def.get("code", ""), sname,
            dims_str, round(cut_len), qty,
            round(unit_wt, 3), round(total_wt, 2),
            loc, etype, added_by, date_added
        ))

    df_main = pd.DataFrame(data, columns=columns)

    # ===================== 2. Summary sheet =====================
    info_data = {
        "Project Name": project_name,
        "Client": client_name,
        "Export Date": datetime.date.today().isoformat()
    }
    df_info = pd.DataFrame(list(info_data.items()), columns=["Item", "Value"])

    if summary_data:
        summary_headers = [
            "Diameter", "Grade", "Total Wt (kg)", "Total Len (m)",
            "# Stock Bars", "Waste (m)", "Waste (%)"
        ]
        df_summary = pd.DataFrame(summary_data, columns=summary_headers)
    else:
        df_summary = pd.DataFrame({"Message": ["No summary data available"]})

    # ===================== 3. Shape Catalog =====================
    shape_rows = []
    for full_name, info in sorted(default_shape_registry.flat_shapes.items()):
        shape_rows.append({
            "Shape Name": full_name,
            "Code": info.get("code", ""),
            "Parameters": ", ".join(info.get("params", [])),
            "Drawing Function": info.get("draw_func", "N/A")
        })
    df_shapes = pd.DataFrame(shape_rows)

    # ===================== Write Excel ==========================
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # --- BBS sheet ---
        df_main.to_excel(writer, sheet_name='Bar Bending Schedule', index=False, startrow=0)
        ws_bbs = writer.sheets['Bar Bending Schedule']
        _format_sheet(ws_bbs, df_main, header_fill=HEADER_FILL, header_font=HEADER_FONT,
                      col_widths={
                          'A': 6, 'B': 14, 'C': 20, 'D': 8, 'E': 8,
                          'F': 8, 'G': 10, 'H': 28, 'I': 22, 'J': 14,
                          'K': 8, 'L': 16, 'M': 14, 'N': 14, 'O': 14,
                          'P': 14, 'Q': 14
                      })

        # --- Summary sheet ---
        start = 0
        df_info.to_excel(writer, sheet_name='Summary', index=False, startrow=start)
        ws_sum = writer.sheets['Summary']
        _format_sheet(ws_sum, df_info, header_fill=HEADER_FILL, header_font=HEADER_FONT,
                      col_widths={'A': 25, 'B': 40})
        start += len(df_info) + 2
        if not df_summary.empty:
            df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=start)
            # Format summary table header
            for col_idx, col_name in enumerate(df_summary.columns, 1):
                cell = ws_sum.cell(row=start + 1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
            _auto_width(ws_sum, df_summary, start + 1)

        # --- Shape Details sheet ---
        if not df_shapes.empty:
            df_shapes.to_excel(writer, sheet_name='Shape Details', index=False)
            ws_shapes = writer.sheets['Shape Details']
            _format_sheet(ws_shapes, df_shapes, header_fill=HEADER_FILL, header_font=HEADER_FONT)

        # --- About (branding) sheet ---
        _write_about_sheet(writer)

    logger.info("Excel export to '%s' completed.", filepath)


# ----------------------------------------------------------------------
# Formatting helpers
# ----------------------------------------------------------------------
def _format_sheet(ws, df, header_fill=None, header_font=None, col_widths=None):
    """Apply standard formatting: bold header, borders, column widths."""
    # Header style
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        if header_fill:
            cell.fill = header_fill
        if header_font:
            cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(df.columns)):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'

    # Column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        _auto_width(ws, df)

    ws.sheet_view.autoFilter = True   # Add auto‑filter
    ws.freeze_panes = 'A2'           # Freeze header row


def _auto_width(ws, df, start_row=1):
    """Guess column widths based on content."""
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name)) + 4
        for row_idx in range(start_row + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)


# ----------------------------------------------------------------------
# About / Branding sheet
# ----------------------------------------------------------------------
def _write_about_sheet(writer):
    """Create an 'About' sheet with company info and optional logo."""
    wb = writer.book
    ws = wb.create_sheet("About")

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = COMPANY_NAME
    ws['A1'].font = Font(name="Calibri", size=22, bold=True, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = COMPANY_TAGLINE
    ws['A2'].font = Font(name="Calibri", size=12, italic=True, color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    # Contact info
    info = [
        ("Website", COMPANY_WEBSITE),
        ("Phone", COMPANY_PHONE),
        ("", ""),
        ("", "© 2025 AI Rebar. All rights reserved."),
        ("", "This report was generated by AI Rebar – Smart Reinforcement Detailing & BBS Software."),
        ("", "For more information and updates, visit our website."),
    ]
    row = 4
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="1F4E79")
        ws.cell(row=row, column=2, value=value).font = Font(color="333333")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

    # Insert logo if available
    if os.path.exists(LOGO_PATH):
        try:
            img = XlImage(LOGO_PATH)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'D1')
        except Exception as e:
            logger.warning("Could not insert logo: %s", e)

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20