# utils/stock_export.py
"""
Export stock inventory report to a professionally formatted Excel file
with bar chart visualization and low‑stock warnings.
"""

import pandas as pd
import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from db.models import StockModel
from config import DEFAULT_REBAR_GRADE

logger = logging.getLogger(__name__)

LOW_STOCK_THRESHOLD = 5          # quantity below which a warning is shown
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", size=10, color="9C0006")


def export_stock_excel(filepath: str, project_id=None) -> None:
    """
    Save the current stock inventory to an Excel workbook with:
      - Inventory sheet (with Status column and conditional formatting)
      - Summary sheet (aggregated totals)
      - Bar chart of total length per diameter
    """
    try:
        stock_items = StockModel.get_all(project_id=project_id)
    except Exception as e:
        logger.error("Failed to fetch stock data: %s", e)
        raise

    columns = ["ID", "Project", "Diameter (mm)", "Grade", "Length (mm)", "Quantity", "Status"]
    data = []
    for row in stock_items:
        rid, pid, dia, length, qty, grade = row
        if grade is None:
            grade = DEFAULT_REBAR_GRADE
        project_label = "Global" if pid is None else f"Project {pid}"
        status = "Low Stock" if qty < LOW_STOCK_THRESHOLD else "OK"
        data.append((rid, project_label, dia, grade, length, qty, status))

    df = pd.DataFrame(data, columns=columns)
    if not df.empty:
        df["Total Length (mm)"] = df["Length (mm)"] * df["Quantity"]
    else:
        df["Total Length (mm)"] = []

    # ---- Write Excel ----
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Inventory sheet
        df.to_excel(writer, sheet_name='Stock Inventory', index=False, startrow=0)
        ws_inv = writer.sheets['Stock Inventory']

        # Apply formatting
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws_inv.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        for row in ws_inv.iter_rows(min_row=2, max_row=ws_inv.max_row, max_col=len(df.columns)):
            for cell in row:
                cell.border = THIN_BORDER
                cell.font = Font(name="Calibri", size=10)

        # Auto-width
        for col_cells in ws_inv.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells) + 2
            ws_inv.column_dimensions[col_cells[0].column_letter].width = min(max_len, 35)

        # Conditional formatting: red fill for "Low Stock"
        status_col = get_column_letter(list(df.columns).index("Status") + 1)
        ws_inv.conditional_formatting.add(
            f"{status_col}2:{status_col}10000",
            CellIsRule(operator='equal', formula=['"Low Stock"'], fill=RED_FILL, font=RED_FONT)
        )

        # Summary sheet
        if not df.empty:
            summary = df.groupby(['Diameter (mm)', 'Grade']).agg(
                Total_Quantity=('Quantity', 'sum'),
                Total_Length_m=('Total Length (mm)', lambda x: x.sum() / 1000),
                No_of_Entries=('Quantity', 'count')
            ).reset_index()
            summary.rename(columns={'Total_Length_m': 'Total Length (m)'}, inplace=True)
            summary.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
            ws_sum = writer.sheets['Summary']
            for col_idx in range(1, len(summary.columns) + 1):
                cell = ws_sum.cell(row=1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
            for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, max_col=len(summary.columns)):
                for cell in row:
                    cell.border = THIN_BORDER
                    if isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'

            # ---- Bar chart: total length per diameter ----
            # Build chart data in a helper sheet
            chart_data = df.groupby('Diameter (mm)')['Total Length (mm)'].sum().reset_index()
            chart_data.to_excel(writer, sheet_name='ChartData', index=False, startrow=0)
            ws_chart = writer.sheets['ChartData']

            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Total Stock Length by Diameter"
            chart.y_axis.title = "Length (mm)"
            chart.x_axis.title = "Diameter (mm)"
            chart.legend = None

            data_ref = Reference(ws_chart, min_col=2, min_row=1, max_row=len(chart_data)+1)
            cats_ref = Reference(ws_chart, min_col=1, min_row=2, max_row=len(chart_data)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 18   # cm
            chart.height = 12  # cm

            # Place chart in the Summary sheet below the table
            ws_sum.add_chart(chart, "A" + str(len(summary) + 3))
        else:
            pd.DataFrame({"Info": ["No stock data available."]}).to_excel(
                writer, sheet_name='Summary', index=False
            )

    logger.info("Stock Excel report saved to '%s'.", filepath)